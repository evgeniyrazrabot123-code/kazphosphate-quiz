import os
import sys
import json
import shutil
import csv
import io
import hashlib
import hmac
from pathlib import Path
from uuid import uuid4
from datetime import date, datetime, timedelta
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from fastapi.responses import StreamingResponse

from fastapi import FastAPI, Depends, UploadFile, File, Form, Response, HTTPException, Header, status, Request, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import Boolean, Float, inspect, Integer, String, Text, DateTime, text
from sqlalchemy.orm import Session
from pydantic import BaseModel

# =====================================================================
# НАСТРОЙКА ПУТЕЙ
# =====================================================================
BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR.parent / "frontend"
UPLOAD_DIR = BASE_DIR / "uploads"

if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

os.makedirs(UPLOAD_DIR, exist_ok=True)

# Импорт моделей и БД
try:
    from .seed_data import seed_questions
except ImportError:
    try:
        from seed_data import seed_questions
    except ImportError:
        seed_questions = None

try:
    from .seed_data import CATEGORIES as START_CATEGORIES
except Exception:
    try:
        from seed_data import CATEGORIES as START_CATEGORIES
    except Exception:
        START_CATEGORIES = []

try:
    from . import models
    from .database import engine, get_db, DATABASE_PATH
except ImportError:
    import models
    from database import engine, get_db, DATABASE_PATH

# =====================================================================
# АВТОРИЗАЦИЯ АДМИНА
# =====================================================================
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'kazphosphate'
ADMIN_TOKEN = 'kazphosphate-admin-token'

def verify_admin_credentials(username: str, password: str) -> bool:
    return username == ADMIN_USERNAME and password == ADMIN_PASSWORD

def get_admin_authorization(authorization: str | None = Header(None)) -> str:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Unauthorized')
    token = authorization.split(' ', 1)[1].strip()
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Unauthorized')
    return token

# =====================================================================
# ИНИЦИАЛИЗАЦИЯ БАЗЫ ДАННЫХ И МИГРАЦИИ
# =====================================================================
def add_missing_columns(engine):
    inspector = inspect(engine)
    existing_tables = set(inspector.get_table_names())

    for table_name, table_obj in models.Base.metadata.tables.items():
        if table_name not in existing_tables:
            continue

        existing_columns = {column_info["name"] for column_info in inspector.get_columns(table_name)}
        for column in table_obj.columns:
            if column.name in existing_columns or column.primary_key:
                continue

            column_type = column.type.compile(engine.dialect)
            add_sql = f'ALTER TABLE "{table_name}" ADD COLUMN "{column.name}" {column_type} NULL'
            if column.server_default is not None:
                default_value = str(column.server_default.arg)
                add_sql = f'ALTER TABLE "{table_name}" ADD COLUMN "{column.name}" {column_type} DEFAULT {default_value} NULL'

            with engine.begin() as conn:
                conn.execute(text(add_sql))

models.Base.metadata.create_all(bind=engine)

# Создаём резервную копию файла БД перед применением ALTER TABLE
try:
    db_path = Path(DATABASE_PATH)
    if db_path.exists():
        bak = db_path.with_suffix(f".bak.{datetime.utcnow().strftime('%Y%m%d%H%M%S')}")
        shutil.copy2(db_path, bak)
except Exception as be:
    print(f"Не удалось создать резервную копию БД: {be}")

add_missing_columns(engine)

# =====================================================================
# ИНИЦИАЛИЗАЦИЯ FASTAPI И MIDDLEWARE
# =====================================================================
app = FastAPI(title="ТОО Казфосфат - Проверка знаний")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# АВТОМАТИЧЕСКАЯ ИНИЦИАЛИЗАЦИЯ БАЗЫ ВОПРОСОВ
@app.on_event("startup")
def startup_event():
    try:
        db = next(get_db())
        if seed_questions and db.query(models.Question).count() == 0:
            try:
                seed_questions()
            except TypeError:
                # backward compatibility: if seed_questions expects a session
                try:
                    seed_questions(db)
                except Exception as e:
                    print(f"Ошибка при запуске seed_questions: {e}")
        # Seed specialties from CATEGORIES if table empty
        try:
            if hasattr(models, 'Specialty') and db.query(models.Specialty).count() == 0 and START_CATEGORIES:
                for code, name_ru, name_kk in START_CATEGORIES:
                    s = models.Specialty(code=code, name_ru=name_ru, name_kk=name_kk)
                    db.add(s)
                db.commit()
        except Exception as se:
            db.rollback()
            print(f"Ошибка при заполнении специальностей: {se}")
    except Exception as e:
        print(f"Ошибка при заполнении начальных данных: {e}")

# Подключение статических файлов (Загрузки и Фронтенд)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

if (FRONTEND_DIR / "js").exists():
    app.mount("/js", StaticFiles(directory=str(FRONTEND_DIR / "js")), name="js")
if (FRONTEND_DIR / "css").exists():
    app.mount("/css", StaticFiles(directory=str(FRONTEND_DIR / "css")), name="css")
if (FRONTEND_DIR / "images").exists():
    app.mount("/images", StaticFiles(directory=str(FRONTEND_DIR / "images")), name="images")

# Serve PWA assets from frontend root (manifest and service worker)
@app.get('/manifest.json')
async def manifest():
    mf = FRONTEND_DIR / 'manifest.json'
    if not mf.exists():
        raise HTTPException(status_code=404, detail='manifest not found')
    return FileResponse(mf, media_type='application/manifest+json')


@app.get('/service-worker.js')
async def service_worker():
    sw = FRONTEND_DIR / 'service-worker.js'
    if not sw.exists():
        raise HTTPException(status_code=404, detail='service-worker not found')
    return FileResponse(sw, media_type='application/javascript')

# =====================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =====================================================================
def save_upload_file(upload_file: UploadFile | None) -> str | None:
    if not upload_file or not getattr(upload_file, "filename", None):
        return None

    file_ext = Path(upload_file.filename).suffix or ".jpg"
    file_name = f"{uuid4().hex}{file_ext}"
    file_path = UPLOAD_DIR / file_name
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)
    return str(file_path).replace(str(BASE_DIR) + os.sep, "")

def safe_json_loads(val):
    if isinstance(val, (list, dict)):
        return val
    try:
        return json.loads(val)
    except Exception:
        return []

# =====================================================================
# РАЗДАЧА HTML СТРАНИЦ И МЕДИА
# =====================================================================
@app.get("/")
async def serve_index():
    index_file = FRONTEND_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="Файл index.html не найден")
    return FileResponse(index_file)

@app.get("/admin")
async def serve_admin():
    admin_file = FRONTEND_DIR / "admin.html"
    if not admin_file.exists():
        raise HTTPException(status_code=404, detail="Файл admin.html не найден")
    return FileResponse(admin_file)

@app.get("/vid.mp4")
async def serve_video():
    video_path = FRONTEND_DIR / "vid.mp4"
    if not video_path.exists():
        raise HTTPException(status_code=404, detail="Видеофайл не найден")
    return FileResponse(video_path, media_type="video/mp4")

# =====================================================================
# API: ПОЛЬЗОВАТЕЛЬСКАЯ ЧАСТЬ
# =====================================================================
@app.get("/api/questions")
def get_questions(category: str, lang: str = "ru", db: Session = Depends(get_db)):
    questions = db.query(models.Question).filter(models.Question.category == category).all()
    
    result = []
    for q in questions:
        raw_opts = q.options_ru if lang == "ru" else q.options_kk
        options = safe_json_loads(raw_opts)
        text_content = q.text_ru if lang == "ru" else q.text_kk
        result.append({
            "id": q.id,
            "text": text_content,
            "options": options
        })
    return result


def hash_employee_password(password: str) -> str:
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 120000)
    return f'{salt.hex()}:{digest.hex()}'


def verify_employee_password(password: str, stored_hash: str | None) -> bool:
    try:
        salt_hex, digest_hex = (stored_hash or '').split(':', 1)
        expected = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), bytes.fromhex(salt_hex), 120000)
        return hmac.compare_digest(expected.hex(), digest_hex)
    except (ValueError, TypeError):
        return False


def find_employee_by_phone(phone: str, db: Session):
    normalized_phone = (phone or '').strip()
    if not normalized_phone:
        return None
    return db.query(models.Employee).filter(models.Employee.phone == normalized_phone).order_by(models.Employee.id.desc()).first()


@app.post('/api/employee/register')
def employee_register(phone: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    normalized_phone = phone.strip()
    if len(normalized_phone) < 7 or len(password) < 6:
        raise HTTPException(status_code=400, detail='Телефон должен содержать минимум 7 символов, пароль — минимум 6.')
    employee = find_employee_by_phone(normalized_phone, db)
    if employee and employee.password_hash:
        raise HTTPException(status_code=409, detail='Этот номер уже зарегистрирован. Выполните вход.')
    if not employee:
        employee = models.Employee(phone=normalized_phone, password_hash=hash_employee_password(password))
        db.add(employee)
    else:
        employee.password_hash = hash_employee_password(password)
    db.commit()
    db.refresh(employee)
    return {"employee_id": employee.id, "full_name": employee.full_name or ''}


@app.post('/api/employee/login')
def employee_login(phone: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    employee = find_employee_by_phone(phone, db)
    if not employee or not verify_employee_password(password, employee.password_hash):
        raise HTTPException(status_code=401, detail='Неверный номер телефона или пароль.')
    return {"employee_id": employee.id, "full_name": employee.full_name}


@app.post('/api/employee/cabinet')
def employee_cabinet(phone: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    employee = find_employee_by_phone(phone, db)
    if not employee or not verify_employee_password(password, employee.password_hash):
        raise HTTPException(status_code=401, detail='Неверный номер телефона или пароль.')

    assignments = db.query(models.TestAssignment).filter(
        models.TestAssignment.employee_id == employee.id
    ).order_by(models.TestAssignment.assigned_at.desc()).all()
    results = db.query(models.TestResult).filter(
        models.TestResult.employee_id == employee.id,
        models.TestResult.is_deleted == False
    ).order_by(models.TestResult.passed_at.desc()).all()

    return {
        "employee": {
            "id": employee.id,
            "full_name": employee.full_name,
            "phone": employee.phone,
            "birth_date": employee.birth_date,
            "position": employee.position,
        },
        "assignments": [{
            "id": assignment.id,
            "category": assignment.category,
            "status": assignment.status,
            "assigned_at": assignment.assigned_at.strftime('%d.%m.%Y %H:%M') if assignment.assigned_at else '—',
        } for assignment in assignments],
        "results": [{
            "id": result.id,
            "score": result.score,
            "total_questions": result.total_questions,
            "passed_at": result.passed_at.strftime('%d.%m.%Y %H:%M') if result.passed_at else '—',
            "passed": result.total_questions > 0 and result.score / result.total_questions >= 0.7,
        } for result in results]
    }


@app.put('/api/employee/profile')
def update_employee_profile(
    phone: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(...),
    birth_date: str = Form(...),
    position: str = Form(...),
    db: Session = Depends(get_db)
):
    employee = find_employee_by_phone(phone, db)
    if not employee or not verify_employee_password(password, employee.password_hash):
        raise HTTPException(status_code=401, detail='Неверный номер телефона или пароль.')
    employee.full_name = full_name.strip()
    employee.birth_date = birth_date.strip()
    employee.position = position.strip()
    db.commit()
    return {"status": "success"}

@app.post("/api/submit")
async def submit_quiz(
    full_name: str = Form(...),
    birth_date: str = Form(None),
    position: str = Form(...),
    iin: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    citizenship: Optional[str] = Form(None),
    employee_password: Optional[str] = Form(None),
    answers: str = Form(...),
    photo_user: UploadFile = File(None),
    photo_license: UploadFile = File(None),
    photo_id_card: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    try:
        # 1. Сохранение всех 3 типов сканов
        user_photo_path = save_upload_file(photo_user)
        lic_photo_path = save_upload_file(photo_license)
        id_card_photo_path = save_upload_file(photo_id_card)

        # 2. Создание карточки сотрудника
        employee = find_employee_by_phone(phone, db)
        if employee and employee.password_hash:
            if not employee_password or not verify_employee_password(employee_password, employee.password_hash):
                raise HTTPException(status_code=401, detail='Необходимо войти в личный кабинет.')
            active_assignment = db.query(models.TestAssignment).filter(
                models.TestAssignment.employee_id == employee.id,
                models.TestAssignment.category == position,
                models.TestAssignment.status == 'assigned'
            ).first()
            if not active_assignment:
                raise HTTPException(status_code=403, detail='Для этого сотрудника тест ещё не назначен руководителем.')
        if employee:
            employee.full_name = full_name.strip() if full_name else employee.full_name
            employee.birth_date = birth_date or employee.birth_date
            employee.position = position
            if user_photo_path:
                employee.photo_user_path = user_photo_path
            if lic_photo_path:
                employee.photo_license_path = lic_photo_path
            if id_card_photo_path:
                employee.photo_id_card_path = id_card_photo_path
        else:
            employee = models.Employee(
                full_name=full_name.strip() if full_name else "",
                birth_date=birth_date,
                position=position,
                iin=iin.strip() if iin and iin.strip() else None,
                phone=phone.strip() if phone else None,
                citizenship=citizenship.strip() if citizenship else None,
                photo_user_path=user_photo_path,
                photo_license_path=lic_photo_path,
                photo_id_card_path=id_card_photo_path,
            )
            db.add(employee)
        db.commit()
        db.refresh(employee)

        # 3. Парсинг и сверка ответов
        answers_dict = json.loads(answers) if isinstance(answers, str) else answers
        score = 0
        details = []

        for q_id_str, user_ans in answers_dict.items():
            q = db.query(models.Question).filter(models.Question.id == int(q_id_str)).first()
            if q:
                user_ans_idx = int(user_ans)
                opts_ru = safe_json_loads(q.options_ru)
                opts_kk = safe_json_loads(q.options_kk)
                
                is_correct = (q.correct_option_index == user_ans_idx)
                if is_correct:
                    score += 1

                details.append({
                    "question_id": q.id,
                    "text_ru": q.text_ru,
                    "text_kk": q.text_kk,
                    "user_answer": user_ans_idx,
                    "correct_answer": q.correct_option_index,
                    "is_correct": is_correct,
                    "options_ru": opts_ru,
                    "options_kk": opts_kk
                })

        total = len(details)

        # 4. Запись результатов в таблицу test_results
        test_result = models.TestResult(
            employee_id=employee.id,
            score=score,
            total_questions=total,
        )
        db.add(test_result)
        assignment = db.query(models.TestAssignment).filter(
            models.TestAssignment.employee_id == employee.id,
            models.TestAssignment.category == position,
            models.TestAssignment.status == 'assigned'
        ).order_by(models.TestAssignment.assigned_at.asc()).first()
        if assignment:
            assignment.status = 'completed'
            assignment.completed_at = datetime.utcnow()
        db.commit()

        return {
            "status": "success", 
            "employee_id": employee.id, 
            "score": score, 
            "total": total,
            "details": details
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ ОШИБКА ПРИ СОХРАНЕНИИ ТЕСТА: {e}")
        raise HTTPException(status_code=500, detail=f"Не удалось сохранить результаты: {str(e)}")

# =====================================================================
# API: АДМИН-ПАНЕЛЬ
# =====================================================================
@app.post('/api/admin/login')
def admin_login(username: str = Form(...), password: str = Form(...)):
    if not verify_admin_credentials(username, password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Invalid credentials')
    return {'token': ADMIN_TOKEN}

@app.get("/api/admin/results")
def get_results_admin(request: Request, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    results = db.query(models.TestResult).filter(getattr(models.TestResult, 'is_deleted', False) == False).all()
    output = []
    
    for r in results:
        emp = r.employee
        
        photo_user_url = None
        photo_license_url = None
        photo_id_card_url = None

        if emp and emp.photo_user_path:
            photo_user_url = str(request.url_for('uploads', path=Path(emp.photo_user_path).name))
        if emp and emp.photo_license_path:
            photo_license_url = str(request.url_for('uploads', path=Path(emp.photo_license_path).name))
        if emp and hasattr(emp, 'photo_id_card_path') and emp.photo_id_card_path:
            photo_id_card_url = str(request.url_for('uploads', path=Path(emp.photo_id_card_path).name))
        
        output.append({
            "id": r.id,
            "employee_id": emp.id if emp else None,
            "passed_at": r.passed_at.strftime("%d.%m.%Y %H:%M") if r and r.passed_at else "—",
            "full_name": emp.full_name if emp else "—",
            "birth_date": emp.birth_date if emp else "—",
            "position": emp.position if emp else "—",
            "iin": getattr(emp, "iin", None),
            "phone": getattr(emp, "phone", None),
            "citizenship": getattr(emp, "citizenship", None),
            "score": r.score,
            "total_questions": r.total_questions,
            "photo_user": photo_user_url,
            "photo_license": photo_license_url,
            "photo_id_card": photo_id_card_url,
        })
        
    return output


@app.post('/api/admin/assignments')
def assign_test(employee_id: int = Form(...), category: str = Form(...), db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail='Сотрудник не найден')
    assignment = models.TestAssignment(employee_id=employee.id, category=category.strip(), status='assigned')
    db.add(assignment)
    db.commit()
    return {"status": "success", "assignment_id": assignment.id}


@app.get('/api/admin/employees')
def get_employees_admin(db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    employees = db.query(models.Employee).order_by(models.Employee.full_name.asc()).all()
    return [{
        "id": employee.id,
        "full_name": employee.full_name or 'Профиль не заполнен',
        "phone": employee.phone or '—',
        "position": employee.position or ''
    } for employee in employees]

@app.delete("/api/admin/results/{r_id}")
def delete_result(r_id: int, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    res = db.query(models.TestResult).filter(models.TestResult.id == r_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Результат не найден")

    # Soft-delete: помечаем результат как удалённый, но не удаляем запись из БД
    if hasattr(res, 'is_deleted'):
        res.is_deleted = True
        db.commit()
        return {"status": "success", "message": "Запись помечена как удалённая (soft-delete)"}

    # Fallback: если нет поля is_deleted, выполняем удаление (старое поведение)
    db.delete(res)
    db.commit()
    return {"status": "success", "message": "Запись успешно удалена"}

@app.get("/api/admin/results/export/csv")
def export_results_csv(request: Request, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    results = db.query(models.TestResult).filter(getattr(models.TestResult, 'is_deleted', False) == False).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header with full employee + result info
    writer.writerow([
        "result_id",
        "employee_id",
        "passed_at",
        "full_name",
        "iin",
        "birth_date",
        "position",
        "citizenship",
        "phone",
        "score",
        "total_questions",
        "percent",
        "photo_user_url",
        "photo_license_url",
        "photo_id_card_url",
        "is_deleted"
    ])

    for r in results:
        emp = r.employee

        photo_user_url = ''
        photo_license_url = ''
        photo_id_card_url = ''
        if emp and emp.photo_user_path:
            photo_user_url = str(request.url_for('uploads', path=Path(emp.photo_user_path).name))
        if emp and emp.photo_license_path:
            photo_license_url = str(request.url_for('uploads', path=Path(emp.photo_license_path).name))
        if emp and getattr(emp, 'photo_id_card_path', None):
            photo_id_card_url = str(request.url_for('uploads', path=Path(emp.photo_id_card_path).name))

        pct = round((r.score / r.total_questions * 100), 1) if r.total_questions > 0 else 0

        writer.writerow([
            r.id,
            emp.id if emp else '',
            r.passed_at.strftime("%Y-%m-%d %H:%M") if r and r.passed_at else '',
            emp.full_name if emp else '',
            emp.iin if emp and getattr(emp, 'iin', None) else '',
            emp.birth_date if emp else '',
            emp.position if emp else '',
            emp.citizenship if emp and getattr(emp, 'citizenship', None) else '',
            emp.phone if emp and getattr(emp, 'phone', None) else '',
            r.score,
            r.total_questions,
            f"{pct}%",
            photo_user_url,
            photo_license_url,
            photo_id_card_url,
            'yes' if getattr(r, 'is_deleted', False) else 'no'
        ])

    response = Response(content=output.getvalue().encode('utf-8-sig'), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=kazphosphate_results.csv"
    return response


@app.get("/api/admin/results/export/xlsx")
def export_results_xlsx(request: Request, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    results = db.query(models.TestResult).filter(getattr(models.TestResult, 'is_deleted', False) == False).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    headers = [
        'result_id', 'employee_id', 'passed_at', 'full_name', 'iin', 'birth_date', 'position', 'citizenship', 'phone',
        'score', 'total_questions', 'percent', 'photo_user_url', 'photo_license_url', 'photo_id_card_url', 'is_deleted'
    ]

    # write header with bold
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for r in results:
        emp = r.employee
        photo_user_url = ''
        photo_license_url = ''
        photo_id_card_url = ''
        if emp and emp.photo_user_path:
            photo_user_url = str(request.url_for('uploads', path=Path(emp.photo_user_path).name))
        if emp and emp.photo_license_path:
            photo_license_url = str(request.url_for('uploads', path=Path(emp.photo_license_path).name))
        if emp and getattr(emp, 'photo_id_card_path', None):
            photo_id_card_url = str(request.url_for('uploads', path=Path(emp.photo_id_card_path).name))

        pct_value = (r.score / r.total_questions) if r.total_questions > 0 else 0

        values = [
            r.id,
            emp.id if emp else None,
            r.passed_at if r and r.passed_at else None,
            emp.full_name if emp else '',
            emp.iin if emp and getattr(emp, 'iin', None) else '',
            emp.birth_date if emp else '',
            emp.position if emp else '',
            emp.citizenship if emp and getattr(emp, 'citizenship', None) else '',
            emp.phone if emp and getattr(emp, 'phone', None) else '',
            r.score,
            r.total_questions,
            pct_value,
            photo_user_url,
            photo_license_url,
            photo_id_card_url,
            'yes' if getattr(r, 'is_deleted', False) else 'no'
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            # percent formatting for the percent column (12th)
            if col_idx == 12:
                cell.number_format = '0.0%'
            # date formatting for passed_at (3rd)
            if col_idx == 3 and isinstance(val, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM'
        row += 1

    # Set column widths for readability
    widths = [12,12,20,30,15,14,25,14,15,8,14,10,40,40,40,10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"kazphosphate_results_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': f'attachment; filename={filename}'
    })

# =====================================================================
# API: СПЕЦИАЛЬНОСТИ
# =====================================================================
@app.get('/api/specialties')
def get_specialties(lang: str = 'ru', db: Session = Depends(get_db)):
    selected_lang = (lang or 'ru').lower()
    try:
        if hasattr(models, 'Specialty'):
            specs = db.query(models.Specialty).all()
            if specs:
                result = {}
                for s in specs:
                    value = s.name_kk if selected_lang == 'kk' and getattr(s, 'name_kk', None) else s.name_ru
                    result[s.code] = value
                return result
    except Exception:
        pass

    # Fallback to START_CATEGORIES
    if START_CATEGORIES:
        result = {}
        for code, name_ru, name_kk in START_CATEGORIES:
            result[code] = name_kk if selected_lang == 'kk' and name_kk else name_ru
        return result
    return {}


@app.get('/api/admin/specialties')
def admin_get_specialties(db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    if not hasattr(models, 'Specialty'):
        return []
    specs = db.query(models.Specialty).all()
    return [{"code": s.code, "name_ru": s.name_ru, "name_kk": s.name_kk} for s in specs]


@app.post('/api/admin/specialties')
def admin_save_specialty(code: str = Form(...), name_ru: str = Form(...), name_kk: str = Form(None), db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    if not hasattr(models, 'Specialty'):
        raise HTTPException(status_code=500, detail='Specialty model not available')

    code_clean = code.strip()
    spec = db.query(models.Specialty).filter(models.Specialty.code == code_clean).first()
    if not spec:
        spec = models.Specialty(code=code_clean, name_ru=name_ru.strip(), name_kk=name_kk)
        db.add(spec)
    else:
        spec.name_ru = name_ru.strip()
        spec.name_kk = name_kk

    db.commit()
    return {"status": "success"}


@app.post('/api/admin/specialties/bulk')
def admin_save_specialties_bulk(specs: dict = Body(...), db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    if not hasattr(models, 'Specialty'):
        raise HTTPException(status_code=500, detail='Specialty model not available')

    try:
        # Upsert each provided specialty
        for code, name in specs.items():
            code_clean = str(code).strip()
            if not code_clean:
                continue
            s = db.query(models.Specialty).filter(models.Specialty.code == code_clean).first()
            if s:
                s.name_ru = str(name)
            else:
                db.add(models.Specialty(code=code_clean, name_ru=str(name)))
        db.commit()
        return {"status": "success"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/api/admin/specialties/{code}')
def admin_delete_specialty(code: str, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    if not hasattr(models, 'Specialty'):
        raise HTTPException(status_code=500, detail='Specialty model not available')
    spec = db.query(models.Specialty).filter(models.Specialty.code == code).first()
    if not spec:
        raise HTTPException(status_code=404, detail='Специальность не найдена')
    db.delete(spec)
    db.commit()
    return {"status": "success"}


# =====================================================================
# API: БЭКАПЫ БД
# =====================================================================
@app.get('/api/admin/db/backups')
def list_db_backups(admin_auth: str = Depends(get_admin_authorization)):
    db_path = Path(DATABASE_PATH)
    if not db_path.exists():
        return []
    parent = db_path.parent
    backups = sorted(parent.glob(db_path.name + '.bak.*'), key=lambda p: p.stat().st_mtime, reverse=True)
    return [{"path": str(p.name), "mtime": p.stat().st_mtime} for p in backups]


@app.post('/api/admin/db/restore')
def restore_db(backup_name: str | None = Form(None), admin_auth: str = Depends(get_admin_authorization)):
    db_path = Path(DATABASE_PATH)
    parent = db_path.parent
    backups = sorted(parent.glob(db_path.name + '.bak.*'), key=lambda p: p.stat().st_mtime, reverse=True)
    if not backups:
        raise HTTPException(status_code=404, detail='Бэкапов не найдено')

    target = None
    if backup_name:
        candidate = parent / backup_name
        if candidate.exists():
            target = candidate
    else:
        target = backups[0]

    if not target:
        raise HTTPException(status_code=404, detail='Указанный бэкап не найден')

    try:
        shutil.copy2(str(target), str(db_path))
        return {"status": "success", "restored": str(target.name)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/questions")
def get_all_questions_admin(db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    questions = db.query(models.Question).all()
    output = []
    for q in questions:
        output.append({
            "id": q.id,
            "category": q.category,
            "text_ru": q.text_ru,
            "text_kk": q.text_kk,
            "options_ru": safe_json_loads(q.options_ru),
            "options_kk": safe_json_loads(q.options_kk),
            "correct_option_index": q.correct_option_index
        })
    return output

@app.post("/api/admin/questions")
def add_question(
    category: str = Form(...),
    text_ru: str = Form(...),
    text_kk: str = Form(...),
    options_ru: str = Form(...),
    options_kk: str = Form(...),
    correct_option_index: int = Form(...),
    db: Session = Depends(get_db),
    admin_auth: str = Depends(get_admin_authorization)
):
    category_clean = category.strip()

    new_q = models.Question(
        category=category_clean,
        text_ru=text_ru.strip(),
        text_kk=text_kk.strip(),
        options_ru=options_ru,
        options_kk=options_kk,
        correct_option_index=int(correct_option_index)
    )
    db.add(new_q)
    db.commit()
    db.refresh(new_q)
    return {"status": "success", "question_id": new_q.id}

@app.delete("/api/admin/questions/{q_id}")
def delete_question(q_id: int, db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    q = db.query(models.Question).filter(models.Question.id == q_id).first()
    if not q:
        return {"status": "error", "message": "Вопрос не найден"}
    db.delete(q)
    db.commit()
    return {"status": "success"}

# =====================================================================
# API: КАРТЫ ДОПУСКА
# =====================================================================
def calculate_expiry_date(start_date: date, category: str) -> date:
    """
    A (Постоянный) — 1 год
    B (Временный) — 6 месяцев (180 дней)
    C (Разовый) — 1 день (24 часа)
    """
    cat = (category or "A").upper()
    if cat == 'A':
        try:
            return start_date.replace(year=start_date.year + 1)
        except Exception:
            return start_date + timedelta(days=365)
    elif cat == 'B':
        # approx 6 months = 180 days
        return start_date + timedelta(days=180)
    elif cat == 'C':
        return start_date + timedelta(days=1)
    try:
        return start_date.replace(year=start_date.year + 1)
    except Exception:
        return start_date + timedelta(days=365)

class PassCardUpdateSchema(BaseModel):
    full_name: Optional[str] = None
    iin: Optional[str] = None
    position: Optional[str] = None
    organization: Optional[str] = None
    personnel_number: Optional[str] = None
    work_type: Optional[str] = None
    access_zone: Optional[str] = None
    vehicle_info: Optional[str] = None
    issued_time: Optional[str] = None
    responsible_person: Optional[str] = None
    category: Optional[str] = None
    issue_date: Optional[date] = None

@app.get("/api/passes/{pass_id}")
def get_pass_card(pass_id: int, db: Session = Depends(get_db)):
    card = db.query(models.PassCard).filter(models.PassCard.id == pass_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Пропуск не найден")
    return card

@app.put("/api/passes/{pass_id}")
def update_pass_card(
    pass_id: int, 
    data: PassCardUpdateSchema, 
    db: Session = Depends(get_db),
    admin_auth: str = Depends(get_admin_authorization)
):
    card = db.query(models.PassCard).filter(models.PassCard.id == pass_id).first()
    
    if not card:
        card = models.PassCard(id=pass_id)
        db.add(card)

    update_dict = data.dict(exclude_unset=True)
    for key, value in update_dict.items():
        if key not in ["category", "issue_date"]:
            setattr(card, key, value)

    new_category = data.category if data.category is not None else (card.category or "A")
    new_issue_date = data.issue_date if data.issue_date is not None else (card.issue_date or date.today())

    card.category = new_category
    card.issue_date = new_issue_date
    card.expiry_date = calculate_expiry_date(new_issue_date, new_category)

    db.commit()
    db.refresh(card)

    return {
        "status": "success",
        "message": "Данные пропуска обновлены",
        "card": {
            "id": card.id,
            "full_name": card.full_name,
            "position": card.position,
            "organization": card.organization,
            "personnel_number": card.personnel_number,
            "work_type": card.work_type,
            "access_zone": card.access_zone,
            "vehicle_info": card.vehicle_info,
            "issued_time": card.issued_time,
            "responsible_person": card.responsible_person,
            "category": card.category,
            "issue_date": card.issue_date.strftime("%Y-%m-%d") if card.issue_date else None,
            "expiry_date": card.expiry_date.strftime("%Y-%m-%d") if card.expiry_date else None,
            "photo_url": card.photo_url,
        }
    }


@app.post('/api/admin/passes/{pass_id}/photo')
def upload_pass_photo(pass_id: int, request: Request, photo: UploadFile = File(None), db: Session = Depends(get_db), admin_auth: str = Depends(get_admin_authorization)):
    card = db.query(models.PassCard).filter(models.PassCard.id == pass_id).first()
    if not card:
        raise HTTPException(status_code=404, detail='Пропуск не найден')

    if not photo or not getattr(photo, 'filename', None):
        raise HTTPException(status_code=400, detail='No photo uploaded')

    # save file
    filename = f"pass_{pass_id}_photo_{uuid4().hex}{Path(photo.filename).suffix}"
    target = UPLOAD_DIR / filename
    with open(target, 'wb') as f:
        shutil.copyfileobj(photo.file, f)

    # update DB
    card.photo_url = str(request.url_for('uploads', path=filename))
    db.commit()
    db.refresh(card)

    return { 'status': 'success', 'photo_url': card.photo_url }